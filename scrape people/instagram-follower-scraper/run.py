import os
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import json
import random
import time
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager as CM

# Complete these 2 fields ==================
USERNAME = 'just_for_fun0323'
PASSWORD = 'happy123456'
# ==========================================

TIMEOUT = 15

small_time_list = [1.4, 1.7, 2.1, 2.3, 2.6, 3.0, 3.8, 4.2]
big_time_list = [180, 190.2, 200.6, 195.5, 213.3, 214.6, 220.9]


def scrape(usr, bot):
    bot.get('https://www.instagram.com/{}/'.format(usr))

    time.sleep(3.5)

    a = bot.find_elements_by_xpath(
        '//*[@id="react-root"]/section/main/div/ul/li[3]/a/div/span')
    user_input = int(a[0].text)

    WebDriverWait(bot, TIMEOUT).until(
        EC.presence_of_element_located((
            By.XPATH, '//*[@id="react-root"]/section/main/div/ul/li[3]/a'))).click()

    time.sleep(2)

    print('[Info] - Scraping...')

    users = set()

    for _ in range(round(user_input // 10)):

        ActionChains(bot).send_keys(Keys.END).perform()

        time.sleep(random.choice(small_time_list))

        followers = bot.find_elements_by_xpath(
            '//*[@id="react-root"]/section/main/div/ul/div/li/div/div[1]/div[2]/div[1]/a')

        # Getting url from href attribute
        for i in followers:
            if i.get_attribute('href'):
                users.add(i.get_attribute('href').split("/")[3])
            else:
                continue

    print('[Info] - Saving...')
    print(f'[DONE] - Your followers are saved in {usr}.txt file!')

    with open(f'{usr}.txt', 'a') as file:
        file.write('\n'.join(users) + "\n")


if __name__ == '__main__':

    # #  ========== log in ============
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless")
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    mobile_emulation = {
        "userAgent": "Mozilla/5.0 (Linux; Android 4.2.1; en-us; Nexus 5 Build/JOP40D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/90.0.1025.166 Mobile Safari/535.19"}
    options.add_experimental_option("mobileEmulation", mobile_emulation)

    bot = webdriver.Chrome(executable_path=CM().install(), options=options)
    bot.set_window_size(600, 1000)

    bot.get('https://www.instagram.com/accounts/login/')

    time.sleep(2)

    print("[Info] - Logging in...")

    user_element = WebDriverWait(bot, TIMEOUT).until(
        EC.presence_of_element_located((
            By.XPATH, '//*[@id="loginForm"]/div[1]/div[3]/div/label/input')))

    user_element.send_keys(USERNAME)

    pass_element = WebDriverWait(bot, TIMEOUT).until(
        EC.presence_of_element_located((
            By.XPATH, '//*[@id="loginForm"]/div[1]/div[4]/div/label/input')))

    pass_element.send_keys(PASSWORD)

    login_button = WebDriverWait(bot, TIMEOUT).until(
        EC.presence_of_element_located((
            By.XPATH, '//*[@id="loginForm"]/div[1]/div[6]/button')))

    time.sleep(0.4)

    login_button.click()

    time.sleep(5)

    # =========== import following_new.xslx =========
    wb_new = openpyxl.load_workbook('../following_new.xlsx')
    sheet_names_new = wb_new.get_sheet_names()
    ws_new = wb_new.get_sheet_by_name(sheet_names_new[0])

    # =========== Opening JSON file ========
    f = open('./../filtered_username_athlete.json')
    data = json.load(f)

    # =========== find total number of edges ======
    edge_num = 0
    while True:
        name = ws_new.cell(row=edge_num + 2, column=1).value
        if name == None:
            break
        edge_num += 1
    print("edge_num: ", edge_num)

    # =========== topic people ==========
    singer_list = []
    for d in data:
        if d['username'] in li:
            singer_list.append(d['username'])
            print(d['username'])
    print(len(singer_list))

    # ========== already_have_people and layer 2 people ==========
    already_have_people = []
    li = []
    for i in range(2, edge_num + 2):
        if i < 15143:
            temp = ws_new.cell(row=i, column=2).value
            if temp not in li:
                li.append(temp)

        name = ws_new.cell(row=i, column=1).value

        if name not in already_have_people:
            already_have_people.append(name)
    print('aleardy have len', len(already_have_people))

    # ===========  scrape people ===========

    for usr in singer_list:
        if usr in already_have_people:
            continue
        try:
            scrape(usr, bot)
        except:
            break
