import pickle
from statistics import mean
import louvain
import igraph as ig
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import matplotlib.pyplot as plt
import ast
import numpy as np


def getGraphAndLayout():
    '''
    G is the graph.
    layout is the position of every nodes
    name_dic is the dictionary whose key is IG name and value is 0 ~ 522 
    '''
    with open('graph_layout.pickle', 'rb') as f:
        G = pickle.load(f)
        layout = pickle.load(f)
        name_dic = pickle.load(f)

        return (G, layout, name_dic)


def getCommunityNumList():
    '''
    get a list of community that is consisted of numbers
    '''
    num_list = []
    with open("community.txt", 'r') as f:
        for line in f.readlines():
            if "community " in line:
                continue

            dic = ast.literal_eval(line)
            num_list.append(dic)

    return num_list


def lengthNodeCommunity(target, community_num_list, func):
    '''
    compute the distance of a node to 5 community 
    '''
    user = name_dic[target]

    for i in range(len(community_num_list)):

        if user in community_num_list[i]:
            community_num_list[i].remove(user)

        temp_list = G.get_shortest_paths(user, community_num_list[i])
        length_list = [len(lis) for lis in temp_list]

        print(f"community{i}'s ", func.__name__, ": ", func(length_list)-1)


def getKey(val, my_dict):
    '''
    get a key of dictionary by giving a value
    '''
    for key, value in my_dict.items():
        if val == value:
            return key

    return "key doesn't exist"


def buildNetwork():
    # ----- input excel -------
    wb = openpyxl.load_workbook(
        '../scrape people/following_split_with_exist.xlsx')
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])

    # ----- get community dictionary -------
    community_dic = []
    with open("community2.txt", 'r') as f:
        for line in f.readlines():
            if "community " in line:
                continue

            dic = ast.literal_eval(line)
            community_dic.append(dic)

    for c in range(7):

        name_dic1 = {}
        G = ig.Graph(len(community_dic[c]), directed=True)
        # ------- build network ---------
        for i in range(2, 9551):

            name = ws.cell(row=i, column=1).value
            following = ws.cell(row=i, column=2).value

            if name not in community_dic[c] or following not in community_dic[c]:
                continue

            if name == None or following == None:
                continue

            if name not in name_dic1:
                name_dic1[name] = len(name_dic1)

            if following not in name_dic1:
                name_dic1[following] = len(name_dic1)

            G.add_edges([(name_dic1[name], name_dic1[following])])

        print(f"community{c}'s transtivity: ", G.transitivity_undirected())
        print(f"community{c}'s density: ", G.density())
        print(f"community{c}'s diameter: ", G.diameter())
        print(f"community{c}'s average path length: ", G.average_path_length())
        # print(G.vcount(), G.ecount())
        print('=' * 10)


# global variable
(G, layout, name_dic) = getGraphAndLayout()

with open("igname2name.pickle", 'rb') as f:
    igname2name_dic = pickle.load(f)


if __name__ == "__main__":
    # (G, layout, name_dic) = getGraphAndLayout()

    # community_dic = getCommunityNumList()
    # user = getKey("高嘉瑜", igname2name_dic)
    # lengthNodeCommunity(user, community_dic, max)

    buildNetwork()
