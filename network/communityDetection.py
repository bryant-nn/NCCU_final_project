
import louvain
import igraph as ig
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import matplotlib.pyplot as plt
import pickle

# ----- input excel -------
wb = openpyxl.load_workbook(
    '../scrape people/following_split_with_exist.xlsx')
sheet_names = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheet_names[0])


# G = ig.Graph(523, directed=True)

#  ------- load public people pickle -----------
file = open('igname2name.pickle', 'rb')
data = pickle.load(file)

# ------- build dictionary of all people ---------


def buildNetwork():
    name_dic1 = {}
    G = ig.Graph(523, directed=True)
    # ------- build network ---------
    for i in range(2, 9551):

        name = ws.cell(row=i, column=1).value
        following = ws.cell(row=i, column=2).value

        if name == None or following == None:
            continue

        if name not in name_dic1:
            name_dic1[name] = len(name_dic1)

        if following not in name_dic1:
            name_dic1[following] = len(name_dic1)

        G.add_edges([(name_dic1[name], name_dic1[following])])

    layout = G.layout(layout='auto')
    # ig.plot(G, layout=layout, target='network.png')

    with open('graph_layout.pickle', 'wb') as f:
        pickle.dump(G, f)
        pickle.dump(layout, f)
        pickle.dump(name_dic1, f)

    return (G, layout, name_dic1)
# ------- function to return key for any value -----------


def get_key(val, my_dict):
    for key, value in my_dict.items():
        if val == value:
            return key

    return "key doesn't exist"


#  ------- community detection ------------
def communityDetection(G, layout, name_dic):
    part = louvain.find_partition(G, louvain.ModularityVertexPartition)
    # print(part)

    output_file = open('community.txt', 'w')
    output_file_1 = open('community1.txt', 'w')
    output_file_2 = open('community2.txt', 'w')

    ret_list = []

    for i in range(len(part)):
        output_file_2.write(f'community {i}\n')
        output_file_1.write(f'community {i}\n')
        output_file.write(f'community {i}\n')

        temp_dic = {}
        chinese_name_list = []

        for people in part[i]:
            name = get_key(people, name_dic)
            temp_dic[name] = data[name]

            chinese_name_list.append(data[name])

        output_file_2.write(str(temp_dic))
        output_file_2.write('\n')

        output_file_1.write(str(chinese_name_list))
        output_file_1.write('\n')

        output_file.write(str(part[i]))
        output_file.write('\n')

        ret_list.append(temp_dic)

    output_file_2.close()
    output_file_1.close()
    output_file.close()

    # ig.plot(part, layout=layout, target='community_detection.png')

    return ret_list


def pageRank(G):
    page_rank = G.pagerank()
    pageRank_dic = {}

    for (k, v) in name_dic.items():
        pageRank_dic[k] = page_rank[v]

    pageRank_dic = {k: v for k, v in sorted(
        pageRank_dic.items(), key=lambda item: item[1], reverse=True)}

    return pageRank_dic


if __name__ == "__main__":

    (G, layout, name_dic) = buildNetwork()

    # print(G.vcount(), G.ecount())
    community_list = communityDetection(G, layout, name_dic)

    pageRank_dic = pageRank(G)

    for i in range(len(community_list)):
        # if len(community_list[i]) <= 5:
        #     continue

        num = 0
        temp_list = []
        for (k, v) in pageRank_dic.items():
            if num == 10:
                break

            if k in community_list[i]:
                temp_list.append(community_list[i][k])
                num += 1
        print(temp_list)
        print('\n')
