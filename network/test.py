
import louvain
import igraph as ig
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import matplotlib.pyplot as plt
import pickle

# ----- input excel -------
wb = openpyxl.load_workbook(
    '../scrape people/following_split_with_exist.xlsx')
sheet_names = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheet_names[0])


G = ig.Graph(523, directed=True)

#  ------- load public people pickle -----------
file = open('igname2name.pickle', 'rb')
data = pickle.load(file)

# ------- build dictionary of all people ---------
name_dic = {}
# for (k, v) in data.items():

#     name_dic[k] = len(name_dic)

# print(name_dic)
# ------- build network ---------
for i in range(2, 9463):
    name = ws.cell(row=i, column=1).value
    following = ws.cell(row=i, column=2).value

    if name == None or following == None:
        continue

    if name not in name_dic:
        name_dic[name] = len(name_dic)

    if following not in name_dic:
        name_dic[following] = len(name_dic)

    # try:
    G.add_edges([(name_dic[name], name_dic[following])])
    # except:
    # print(name, following)
    # print(name_dic[name], name_dic[following])
    # break
# ------- function to return key for any value -----------


def get_key(val, my_dict):
    for key, value in my_dict.items():
        if val == value:
            return key

    return "key doesn't exist"


#  ------- community detection ------------
part = louvain.find_partition(G, louvain.ModularityVertexPartition)
# print(part)

output_file = open('community1.txt', 'w')
output_file_1 = open('community.txt', 'w')

for i in range(len(part)):
    output_file.write(f'community {i}\n')
    output_file_1.write(f'community {i}\n')

    temp_list = []
    for people in part[i]:
        name = get_key(people, name_dic)
        temp_list.append(data[name])

    output_file.write(str(temp_list))
    output_file.write('\n\n\n\n')

    output_file_1.write(str(part[i]))
    output_file_1.write('\n\n\n\n')

output_file.close()
output_file_1.close()

ig.plot(part)

print(type(part), G.vcount(), G.ecount())
